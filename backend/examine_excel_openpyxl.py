from openpyxl import load_workbook
import os

def examine_excel_with_openpyxl(file_path):
    """Examine Excel file using openpyxl"""
    try:
        print(f"Reading Excel file: {file_path}")
        
        # Load the workbook
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        
        print(f"Workbook loaded successfully!")
        print(f"Sheets found: {workbook.sheetnames}")
        
        # Examine each sheet
        for sheet_name in workbook.sheetnames:
            print(f"\n{'='*50}")
            print(f"Sheet: {sheet_name}")
            print('='*50)
            
            sheet = workbook[sheet_name]
            
            # Get dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            print(f"Dimensions: {max_row} rows x {max_col} columns")
            
            # Read first few rows to understand structure
            print("\nFirst 10 rows (first 10 columns):")
            for row in range(1, min(11, max_row + 1)):
                row_data = []
                for col in range(1, min(11, max_col + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else '')
                print(f"Row {row}: {row_data}")
            
            # Try to identify headers
            if max_row > 0:
                print(f"\nPotential headers (Row 1):")
                headers = []
                for col in range(1, min(11, max_col + 1)):
                    header_value = sheet.cell(row=1, column=col).value
                    headers.append(str(header_value) if header_value is not None else f'Column_{col}')
                print(headers)
        
        workbook.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Path to the Excel file
    excel_file_path = "../BranchSalesAnalysisDepartment (4).xls"
    
    if os.path.exists(excel_file_path):
        examine_excel_with_openpyxl(excel_file_path)
    else:
        print(f"File not found: {excel_file_path}")
        print("Current directory:", os.getcwd())
        print("Files in current directory:", os.listdir(".")) 